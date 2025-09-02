
from .utils import EXPORTS_DIR
from .db import get_conn
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def _fmt_header(ws, row=1):
    bold = Font(bold=True)
    for cell in ws[row]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

def export_excel():
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet CR
    ws_cr = wb.active
    ws_cr.title = "CR"
    ws_cr.append(["ID","Date","Thématique","Projet","Titre","Participants","Contenu"])
    _fmt_header(ws_cr)

    con = get_conn(); cur = con.cursor()
    cur.execute("SELECT id,date,thematique,projet,titre,participants,content FROM meetings ORDER BY date DESC, id DESC")
    for row in cur.fetchall():
        ws_cr.append(row)

    # Sheet ToDo
    ws_td = wb.create_sheet("ToDo")
    ws_td.append(["ID","Meeting_ID","Thématique","Projet","Action","Acteur","Échéance","Statut"])
    _fmt_header(ws_td)
    cur.execute("SELECT id,meeting_id,thematique,projet,action,acteur,echeance,statut FROM todos ORDER BY id DESC")
    for row in cur.fetchall():
        ws_td.append(row)

    # Sheet ToDo_Global
    ws_tdg = wb.create_sheet("ToDo_Global")
    ws_tdg.append(["ID","Date réunion","Thématique","Projet","Action","Acteur","Échéance","Statut","Meeting_ID","Titre"])
    _fmt_header(ws_tdg)
    cur.execute("""
        SELECT t.id, m.date, t.thematique, t.projet, t.action, t.acteur, t.echeance, t.statut, t.meeting_id, m.titre
        FROM todos t LEFT JOIN meetings m ON t.meeting_id = m.id
        ORDER BY m.date DESC, t.id DESC
    """)
    for row in cur.fetchall():
        ws_tdg.append(row)

    con.close()

    fname = f"CHAP1_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out = EXPORTS_DIR / fname
    wb.save(out)
    return out
