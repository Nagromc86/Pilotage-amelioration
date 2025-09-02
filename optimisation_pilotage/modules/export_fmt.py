
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from . import db, utils
except ImportError:
    # If imported as a standalone for tests
    from optimisation_pilotage.modules import db, utils  # type: ignore

def _sanitize_sheet_name(name: str) -> str:
    invalid = '[]:*?/\\'
    for ch in invalid:
        name = name.replace(ch, ' ')
    name = name.strip()
    return name[:31] if len(name) > 31 else name

def export_cr_formatted_tables(theme_id=None, project_id=None) -> str:
    """
    Crée un classeur Excel avec 1 feuille Index et 1 feuille par réunion,
    contenant un tableau Label / Valeur, pratique pour copier-coller dans Word/PPT.
    Retourne le chemin du fichier Excel généré.
    """
    conn = db.get_conn()
    rows = conn.execute("""
        SELECT m.id, m.theme_id, m.project_id, m.title, m.date, COALESCE(m.participants,''),
               COALESCE(m.summary,''), COALESCE(m.transcript_path,''),
               th.name, p.name
        FROM meetings m
        LEFT JOIN themes th ON th.id = m.theme_id
        LEFT JOIN projects p ON p.id = m.project_id
        WHERE (? IS NULL OR m.theme_id=?)
          AND (? IS NULL OR m.project_id=?)
        ORDER BY m.date DESC, m.id DESC
    """, (theme_id, theme_id, project_id, project_id)).fetchall()

    wb = Workbook()
    ws_idx = wb.active
    ws_idx.title = "Index"
    ws_idx.append(["MeetingID", "Date", "Thématique", "Projet", "Titre", "Feuille"])
    bold = Font(bold=True)
    for c in ws_idx[1]:
        c.font = bold

    head_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill("solid", fgColor="E6EEF9")
    thin = Side(style="thin", color="B7C1D6")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in rows:
        mid, theme_id, project_id, title, date, participants, summary, tpath, theme_name, proj_name = row
        # contenu (fichier transcription si dispo, sinon summary)
        txt = ""
        if tpath and Path(tpath).exists():
            try: txt = Path(tpath).read_text(encoding="utf-8")
            except Exception: pass
        if not txt: txt = summary or ""

        todos = conn.execute("""
            SELECT action, COALESCE(actor,''), COALESCE(due_date,''), status
            FROM todos WHERE meeting_id=? ORDER BY id
        """, (mid,)).fetchall()
        bullets = "\\n".join([f"• {t[0]} (Acteur: {t[1] or '—'}, Échéance: {t[2] or '—'}, Statut: {t[3]})" for t in todos]) or "—"

        sheet_name = _sanitize_sheet_name(f"{date}_{title}")
        # éviter doublons de nom
        base = sheet_name
        n = 2
        while sheet_name in wb.sheetnames:
            sheet_name = _sanitize_sheet_name(f"{base[:28]}_{n}")
            n += 1

        ws = wb.create_sheet(title=sheet_name)

        # Titre
        ws.merge_cells("A1:B1")
        ws["A1"] = f"COMPTE RENDU — {title}"
        ws["A1"].font = head_font

        # En-têtes du tableau
        ws["A3"] = "Élément"; ws["A3"].font = label_font; ws["A3"].fill = header_fill; ws["A3"].border = border_all; ws["A3"].alignment = wrap_top
        ws["B3"] = "Valeur";  ws["B3"].font = label_font; ws["B3"].fill = header_fill; ws["B3"].border = border_all; ws["B3"].alignment = wrap_top

        data_rows = [
            ("Date", date or ""),
            ("Thématique", theme_name or ""),
            ("Projet", proj_name or ""),
            ("Participants", participants or ""),
            ("Synthèse", txt or ""),
            ("Actions", bullets),
        ]

        r = 4
        for label, value in data_rows:
            ws[f"A{r}"] = label
            ws[f"A{r}"].font = label_font
            ws[f"A{r}"].alignment = wrap_top
            ws[f"B{r}"] = value
            ws[f"B{r}"].alignment = wrap_top
            ws[f"A{r}"].border = border_all
            ws[f"B{r}"].border = border_all
            r += 1

        # mise en forme
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 100
        ws.freeze_panes = "A4"

        # Index + lien interne
        ws_idx.append([mid, date or "", theme_name or "", proj_name or "", title or "", sheet_name])
        link_cell = ws_idx.cell(row=ws_idx.max_row, column=6)
        link_cell.hyperlink = f"#{sheet_name}!A1"
        link_cell.style = "Hyperlink"

    # Index format
    ws_idx.column_dimensions["A"].width = 10
    ws_idx.column_dimensions["B"].width = 12
    ws_idx.column_dimensions["C"].width = 24
    ws_idx.column_dimensions["D"].width = 24
    ws_idx.column_dimensions["E"].width = 50
    ws_idx.column_dimensions["F"].width = 26
    ws_idx.freeze_panes = "A2"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(utils.EXPORTS_DIR) / f"CR_formates_tableaux_{ts}.xlsx"
    wb.save(out)
    return str(out)
